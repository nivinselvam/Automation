import openpyxl
from DataProvider import GlobalVariables

matchFound = False
global sheet

def initializeWorkbook(workbookPath):
    workbook = openpyxl.load_workbook(workbookPath)
    return workbook

def initializeSheet(workbook, sheetName):
    sheet = workbook[sheetName]
    return sheet

def getColumnNumberFromName(workbook, sheet, columnName):
    matchFound = False
    totalcolumns = sheet.max_column
    i=1
    for col in sheet.iter_cols(0, totalcolumns):
        if col[0].value == columnName:
            matchFound = True
            break
        i+=1
    if matchFound:
        return i
    else:
        print("Entered column value not available. Please verify the Data sheet.")
        return ValueError

def getRowNumberFromValue(workbook, sheet, columnName, rowValue):
    matchFound = False
    totalrows = sheet.max_row
    columnNumber = getColumnNumberFromName(workbook, sheet, columnName)
    i=1
    for currentRow in range(1, totalrows+1):
        if sheet.cell(row=currentRow,column=columnNumber).value == rowValue:
            matchFound = True
            break
        i+=1
    if matchFound:
        return i
    else:
        print("Entered row value not available. Please verify the Data sheet.")
        return ValueError

def getAPIEnviURL(workbook, environment):
    sheet = initializeSheet(workbook, "Environment")
    columnNumber = getColumnNumberFromName(workbook, sheet, "URL")
    rowNumber = getRowNumberFromValue(workbook, sheet, "Name", environment)

    if columnNumber == ValueError or rowNumber == ValueError:
        print("Environment details unavailable. Please verify the Data sheet.")
        return ValueError
    else:
        return sheet.cell(row=rowNumber, column=columnNumber).value

def getAPIDetails(workbook, apiName, apidetailRequired):
    sheet = initializeSheet(workbook, "APIDetails")
    columnNumber = getColumnNumberFromName(workbook, sheet, apidetailRequired)
    rowNumber = getRowNumberFromValue(workbook, sheet, "Action", apiName)

    if columnNumber == ValueError or rowNumber == ValueError:
        print("API details is  not available. Please verify the Data sheet.")
        return ValueError
    else:
        return sheet.cell(row=rowNumber, column=columnNumber).value
    
def set_cell_value(excel_path, row_value, column_name, value):
    i = 0
    while i <10:
        try:
             with open(excel_path, 'a') as myfile:
                fcntl.flock(myfile, fcntl.LOCK_EX | fcntl.LOCK_NB)
                workbook = openpyxl.load_workbook(filename= excel_path)
                sheet = initializeSheet(workbook, "result")
                column_number = getColumnNumberFromName(workbook, sheet, column_name)
                row_number = getRowNumberFromValue(workbook,sheet,"TestCase", row_value)
                current_cell = sheet.cell(row=row_number, column=column_number)
                current_cell.value = value
                workbook.save(excel_path)
                workbook.close()
                fcntl.flock(myfile, fcntl.LOCK_UN)
                return "Success"
        except Exception as e:
            logger.info(f"Writing to excel failed due to error {str(e)}")
            time.sleep(1)
            i +=1
    return "Failure"


def getAPIValidations(workbook, apiName):
    sheet = initializeSheet(workbook, "APIDetails")
    columnNumber = getColumnNumberFromName(workbook, sheet, "API_Validation")
    rowNumber = getRowNumberFromValue(workbook, sheet, "Action", apiName)

    if columnNumber == ValueError or rowNumber == ValueError:
        print("API details not available. Please verify the Data sheet.")
        return ValueError
    else:
        return sheet.cell(row=rowNumber, column=columnNumber).value

def getDBValidations(workbook, apiName):
    sheet = initializeSheet(workbook, "APIDetails")
    columnNumber = getColumnNumberFromName(workbook, sheet, "DB_Validation")
    rowNumber = getRowNumberFromValue(workbook, sheet, "Action", apiName)

    if columnNumber == ValueError or rowNumber == ValueError:
        print("DB details not available. Please verify the Data sheet.")
        return ValueError
    else:
        return sheet.cell(row=rowNumber, column=columnNumber).value

def getPortalValidations(workbook, apiName):
    sheet = initializeSheet(workbook, "APIDetails")
    columnNumber = getColumnNumberFromName(workbook, sheet, "Portal_Validation")
    rowNumber = getRowNumberFromValue(workbook, sheet, "Action", apiName)

    if columnNumber == ValueError or rowNumber == ValueError:
        print("Portal details not available. Please verify the Data sheet.")
        return ValueError
    else:
        return sheet.cell(row=rowNumber, column=columnNumber).value
